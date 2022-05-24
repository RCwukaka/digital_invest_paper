# coding=gbk
from pymysql_comm import UsingMysql


def create_one(data):
    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into tuobinq_analyzed(tuobinQ,time,ts_code) " + \
                  "values(%s,%s,%s)"
            params = (data['tuobinq'], data['time'], data['ts_code'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)


from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('../data/tuobinq_data.xlsx')
    sheets = wb.worksheets  # 获取当前所有的sheet

    # 获取第一张sheet
    sheet1 = sheets[0]

    for i in range(1, sheet1.max_row):
        tuobinq = str(sheet1.cell(i, 3).value)
        time = sheet1.cell(i, 2).value[5:7]
        if time == '12':
            data = {}
            ts_code = str(sheet1.cell(i, 1).value)
            year = str(sheet1.cell(i, 2).value[:4])
            data['time'] = year
            data['ts_code'] = ts_code
            data['tuobinq'] = tuobinq
            create_one(data)
