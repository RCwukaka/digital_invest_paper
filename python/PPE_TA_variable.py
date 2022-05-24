# coding=gbk
from pymysql_comm import UsingMysql


def create_one(data):
    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into PPE_TA(ppe_ta,time,ts_code) " + \
                  "values(%s,%s,%s)"
            params = (data['ppe_ta'], data['time'], data['ts_code'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)


from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('../data/PPE_TA.xlsx')
    sheets = wb.worksheets  # 获取当前所有的sheet

    # 获取第一张sheet
    sheet1 = sheets[0]

    for i in range(1, sheet1.max_row):
        data = {}
        data['time'] = sheet1.cell(i, 2).value[:4]
        data['ts_code'] = sheet1.cell(i, 1).value
        data['ppe_ta'] = sheet1.cell(i, 3).value
        create_one(data)
