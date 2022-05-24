# coding=gbk
from pymysql_comm import UsingMysql

# 新增单条记录
def create_one(data):

    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into economy_index(economy_val,time,province) " + \
                  "values(%s,%s,%s)"
            params = (data['economy_val'],data['time'],data['province'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)

from openpyxl import load_workbook

if __name__ == '__main__':

    wb = load_workbook('../data/economy_data.xlsx')
    sheets = wb.worksheets  # 获取当前所有的sheet

    # 获取第一张sheet
    sheet1 = sheets[0]

    province = []
    for i in range(2, 33):
        province.append(sheet1.cell(i, 1).value)

    time = []
    for i in range(2, 14):
        time.append(sheet1.cell(1, i).value)

    data = {}
    for i in range(2, 33):
        for j in range(2, 14):
            education = sheet1.cell(i, j).value
            data['economy_val'] = education
            data['time'] = time[j - 2]
            data['province'] = province[i - 2]
            create_one(data)
