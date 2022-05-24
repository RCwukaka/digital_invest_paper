# coding=gbk
from pymysql_comm import UsingMysql


# 新增单条记
def create_one(data):
    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into roa(ts_code,roa_val,time) " + \
                  "values(%s,%s,%s)"
            params = (data['ts_code'], data['roa'], data['time'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)


from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('../data/roa.xlsx')
    sheets = wb.worksheets  # 获取当前所有的sheet

    # 获取第一张sheet
    sheet1 = sheets[0]

    data = {}
    for i in range(1, sheet1.max_row):
        type = sheet1.cell(i, 3).value
        time = sheet1.cell(i, 2).value[6:7]
        if type.find("A") != -1 and time == '12':
            data = {}
            money = 0
            if sheet1.cell(i, 1).value != None and sheet1.cell(i, 4).value != None:
                ts_code = str(sheet1.cell(i, 1).value)
                time = sheet1.cell(i, 2).value[:4]
                data['time'] = time
                data['ts_code'] = ts_code
                data['roa'] = sheet1.cell(i, 4).value
                create_one(data)
