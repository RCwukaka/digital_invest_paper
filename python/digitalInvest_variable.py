#coding=gbk
from pymysql_comm import UsingMysql

def create_one(data):

    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into digitalinvest(money,time,ts_code) " + \
                  "values(%s,%s,%s)"
            params = (data['money'],data['time'],data['ts_code'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)


from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('../data/digitalinvest_data.xlsx')
    sheets = wb.worksheets  # ��ȡ��ǰ���е�sheet

    # ��ȡ��һ��sheet
    sheet1 = sheets[0]

    data = {}
    for i in range(1, sheet1.max_row):
        type = str(sheet1.cell(i,5).value)
        if type.find("���")!=-1 or type.find("����ϵͳ")!=-1 or type.find("����")!=-1 or type.find("����")!=-1 or type.find("����ƽ̨")!=-1 or type.find("�ͻ���")!=-1:
            data={}
            money = 0
            if sheet1.cell(i, 7).value != None and sheet1.cell(i, 6).value != None:
                money = int(sheet1.cell(i, 7).value) - int(sheet1.cell(i, 6).value)
            if money<0:money=0
            ts_code = str(sheet1.cell(i,1).value)
            time = sheet1.cell(i,2).value[:4]
            data['time'] = time
            data['ts_code'] = ts_code
            data['money'] = money
            create_one(data)


    wb = load_workbook('../data/fixed_assets_data.xlsx')
    sheets = wb.worksheets  # ��ȡ��ǰ���е�sheet

    # ��ȡ��һ��sheet
    sheet1 = sheets[0]

    data = {}
    for i in range(1, sheet1.max_row):
        type = str(sheet1.cell(i, 5).value)
        if type.find("�����") != -1 or type.find("������") != -1:
            data = {}
            money = sheet1.cell(i, 9).value
            if money < 0: money = 0
            ts_code = str(sheet1.cell(i, 1).value)
            time = sheet1.cell(i, 2).value[:4]
            data['time'] = time
            data['ts_code'] = ts_code
            data['money'] = money
            create_one(data)