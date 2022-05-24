# coding=gbk

from openpyxl import load_workbook
from pymysql_comm import UsingMysql


# 新增单条记录
def create_one(data):
    with UsingMysql(log_time=True) as um:
        try:
            sql = "insert into education(education_val,time,province) " + \
                  "values(%s,%s,%s)"
            params = (data['education_val'], data['time'], data['province'])
            um.cursor.execute(sql, params)
        except BaseException as e:
            print(e)


def select_by_tscode(ts_code):
    with UsingMysql(log_time=True) as um:
        sql = "SELECT * FROM announcements_list where ts_code = %s"
        params = (ts_code)
        um.cursor.execute(sql, params)
        return um.cursor.fetchall()


def select_all():
    with UsingMysql(log_time=True) as um:
        sql = "SELECT * FROM announcements_list where id>45774"
        um.cursor.execute(sql)
        return um.cursor.fetchall()


def update_by_pk(ann_date, id):
    with UsingMysql(log_time=True) as um:
        sql = "update announcements_list set ts_code = %s where id = %s"
        params = (ann_date, id)
        um.cursor.execute(sql, params)


wb = load_workbook('../data/education_data.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]

province = []
for i in range(2, 33):
    province.append(sheet1.cell(i, 1).value)

time = []
for i in range(2, 20):
    time.append(sheet1.cell(1, i).value)

data = {}
for i in range(2, 33):
    for j in range(2, 20):
        education = sheet1.cell(i, j).value
        data['education_val'] = education
        data['time'] = time[j - 2]
        data['province'] = province[i - 2]
        create_one(data)
